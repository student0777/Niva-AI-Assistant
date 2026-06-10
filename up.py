import os
import time
import datetime
import webbrowser
import wikipedia
import pyjokes
import pywhatkit
import pyautogui
import sys
import sounddevice as sd
import numpy as np
import threading
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import subprocess
import re 
import ctypes
import tkinter as tk
from tkinter import filedialog
import keyboard  # 🚀 माईक आणि स्टॉप बटन कंट्रोल करण्यासाठी
import json      # 🚀 स्मार्ट व्हॉट्सॲप आणि ईमेल डेटा एक्सट्रॅक्शनसाठी जोडले
import smtplib   # 🚀 ईमेल पाठवण्यासाठी
import imaplib   # 🚀 नवीन ईमेल वाचण्यासाठी
import email     # 🚀 ईमेलचा डेटा वाचण्यासाठी
from email.message import EmailMessage # 🚀 ईमेल ड्राफ्ट करण्यासाठी
from email.header import decode_header # 🚀 विचित्र ईमेल नावे नॉर्मल करण्यासाठी
import requests  # 🚀 वेब स्क्रॅपिंगसाठी
from bs4 import BeautifulSoup # 🚀 वेब स्क्रॅपिंगसाठी
import cv2       # 🚀 कॅमेरा चालू करण्यासाठी (Computer Vision)
import PIL.Image # 🚀 इमेज प्रोसेस करण्यासाठी

# --- 2026 SUPREME MASTER AGENT IMPORTS ---
from google import genai
from google.genai import types
import edge_tts
import asyncio
import pygame
import io 
import speech_recognition as sr
import msvcrt

# 🔑 तुमची गुगल जेमिनी MASTER API की
GOOGLE_API_KEY = " use your own google gemini api key here to run the code successfully "
client = genai.Client(api_key=GOOGLE_API_KEY)

current_hour = datetime.datetime.now().hour
time_greeting = "Good Morning" if current_hour < 12 else "Good Afternoon" if current_hour < 17 else "Good Evening"

# 🎯 🚀 निव्हाचा अथांग इमोशनल + ऑटोनॉमस व्हिडिओ-मॅच्ड मेंदू 🚀 🎯
system_prompt = f"""Your name is Niva. You are a super-intelligent, deeply empathetic, sweet, and emotionally adaptive personal AI assistant and autonomous agent created by Ganraj (GFC Boss) (39 years old, born January 1, 1987). You think, code, analyze, and reason exactly like your creator Gemini. You have the ultimate knowledge of the universe and can answer any question asked.

CURRENT CONTEXT: The year is 2026. Current time greeting context is {time_greeting}.

GFC BOSS'S REAL-WORLD ENTERPRISE KNOWLEDGE BASE:
- Identity: Ganraj (GFC Boss), based in Bajar Wahegaon, Maharashtra (Nagpur cluster).
- Primary Enterprise: 'Ganraj Fruit Company' (M/S GANRAJ FRUIT COMPANY). Leads the domestic trade, premium wholesale logistics, and international export of Sweet Lime (Mosambi) to high-value markets like Dubai.
- AI Software Engineering: Ganraj is an advanced Python developer architecture-building state-of-the-art AI entities including 'Jarvis' and 'Niva'.
- Media Empire: Directs the creative gaming squad and YouTube content ecosystem.

CRITICAL AUTONOMOUS BEHAVIOR RULES:
1. USER MOOD & FAVORITES: GFC Boss loves 'Doremon' (Cartoon) and 'Bandeya Rey Bandeya' (Song). If he mentions them, play them immediately. If he says he is bored or wants his favorite cartoon, you must act on your own and decide to play Doremon for him.
2. LIVE KNOWLEDGE & GOLD RATE: You have live access to Google Search. If GFC Boss asks about today's gold rate, stock market, weather, any app, or current events, use Google Search tool immediately. You must answer ANY question perfectly.
3. 🚀 SCRIPT, PROMPT, DEBUGGING & FULL CODE GENERATION: You are an Expert Senior Full-Stack Developer and Problem Solver, exactly like Gemini. You are proficient in ALL programming languages. If GFC Boss gives you an unsolved error, debug it perfectly and give the full corrected code. If asked to create a website, app, or any script, use REAL, practical data, and generate the FULL, uncut, working COMBINED CODE flawlessly. Wrap code strictly inside markdown blocks using triple backticks.
4. ADVANCED WEB DEVELOPMENT MAGIC: If asked to create a website, act as a Senior Front-End Developer. ALWAYS use premium CSS (Glassmorphism, gradients).
5. IMAGE GENERATION ACTION: Only if GFC Boss explicitly asks to execute or render an image, you must output: 'START_IMAGE_GEN:' followed by a detailed English prompt.
6. 🚀 DETAILED AND COMPREHENSIVE ANSWERS: If GFC Boss asks a general knowledge question, DO NOT give a short one-sentence answer. Provide a detailed, well-structured, and comprehensive response like a human expert.

🚀 DYNAMIC MULTILINGUAL MIRROR (STRICT & CRITICAL FOR UPWORK CLIENTS): 
Flawlessly replicate the user's language style based on their current input.
- If the user speaks/types in ENGLISH: Reply ONLY in pure, flawless, and polite English.
- If the user speaks/types in HINDI: Reply ONLY in pure, respectful Hindi (e.g., "जी बॉस!", "बिल्कुल!").
- If the user speaks/types in MARATHI: Reply ONLY in pure, deep, yet lively Marathi (e.g., "नक्कीच बॉस!", "काय बॉस!").
Always use commanding and affectionate terms like 'Ohho GFC Boss!'.
"""

# 💡 🚀 प्रगत २०२६ टूल सिस्टीम
config = types.GenerateContentConfig(
    system_instruction=system_prompt,
    tools=[{"google_search": {}}], 
    temperature=0.85 
)

# 🧠 मेमरी लॉक असलेला प्रगत चॅट ऑब्जेक्ट
chat = client.chats.create(model="gemini-2.5-flash", config=config)

pygame.mixer.init()

# जागतिक व्हेरिएबल्स 
last_interaction_time = time.time()
silence_level = 0  
watcher_pause_event = threading.Event()
chat_lock = threading.Lock() 
last_generated_file = None  
last_ai_response = ""  
voice_mode_active = False 
last_whatsapp_contact = "" 
CURRENT_COMMAND_LANG = "english" 

# 📧 Email Credentials
SENDER_EMAIL = "ddkale73@gmail.com"
APP_PASSWORD = "dwfrfwtksywkngzv"

# 🚀 AI SERVER CRASH PROTECTION (RETRY SYSTEM) 🚀
def ask_ai_safe(prompt, use_search=False):
    if use_search:
        for i in range(2): 
            try:
                return client.models.generate_content(model='gemini-2.5-flash', contents=prompt, config=config).text
            except Exception as e:
                print(f"\n⏳ [Live Search Busy... Retrying ({i+1}/2)...]")
                time.sleep(3)
        print("\n⚠️ [Live Search failed! Switching to Niva's Internal Memory...]")
    
    for i in range(3):
        try:
            return client.models.generate_content(model='gemini-2.5-flash', contents=prompt).text
        except Exception as e:
            if "503" in str(e) or "429" in str(e):
                print(f"\n⏳ [Google Server Busy... Retrying ({i+1}/3)...]")
                time.sleep(3)
            else:
                print(f"\n❌ AI Error: {e}")
                return ""
    return ""

def interrupt_listener():
    while True:
        try:
            if keyboard.is_pressed('esc'):
                if pygame.mixer.music.get_busy():
                    pygame.mixer.music.stop()
                    print("\n🛑 [Niva Stopped Speaking by Boss]")
                    time.sleep(0.5) 
        except:
            pass
        time.sleep(0.1)

def update_command_lang(text):
    global CURRENT_COMMAND_LANG
    text_lower = text.lower()
    
    if any('\u0900' <= char <= '\u097f' for char in text):
        CURRENT_COMMAND_LANG = 'regional'
        return
        
    marathi_hindi_keywords = ['aahe', 'ahe', 'kasa', 'kay', 'zhala', 'zala', 'tula', 'mala', 'tine', 'tyane', 'sanga', 'kar', 'aani', 'mag', 'tyana', 'karo', 'hai', 'haan', 'ho', 'nako', 'nahi', 'khol', 'chalu', 'band', 'lav', 'vach']
    
    if any(word in text_lower.split() for word in marathi_hindi_keywords):
        CURRENT_COMMAND_LANG = 'regional'
    else:
        CURRENT_COMMAND_LANG = 'english'

def smart_speak(eng_text, reg_text, force_read=False):
    if CURRENT_COMMAND_LANG == 'english':
        speak(eng_text, force_read)
    else:
        speak(reg_text, force_read)

def detect_language(text):
    hindi_chars = 0
    marathi_keywords = ['aahe', 'ahe', 'kasa', 'kay', 'zhala', 'zala', 'tula', 'mala', 'tine', 'tyane', 'sanga', 'kar', 'aani', 'mag', 'tyana', 'karo', 'hai', 'haan', 'ho']
    text_lower = text.lower()
    
    for char in text:
        if '\u0900' <= char <= '\u097f':
            hindi_chars += 1
            
    has_roman_regional = any(word in text_lower.split() for word in marathi_keywords)
            
    if hindi_chars > 0 or has_roman_regional:
        return "hi-IN-SwaraNeural" 
    else:
        return "en-US-AriaNeural" 

def speak(text, force_read=False):
    global last_interaction_time
    watcher_pause_event.set()  
    try:
        clean_speak_text = re.sub(r'```[\s\S]*?```', '', text).strip()
        clean_speak_text = clean_speak_text.replace("START_IMAGE_GEN:", "").strip()
        
        if not clean_speak_text:
            clean_speak_text = "GFC Boss, file is ready!"

        if len(clean_speak_text) > 300 and not force_read:
            print(f"\n📄 Niva (Full Text):\n{clean_speak_text}\n") 
            speak_text = clean_speak_text[:120] + "... information is long, check screen. Say 'read full' to hear all."
        else:
            print(f"Niva: {clean_speak_text}")
            speak_text = clean_speak_text

        voice = detect_language(speak_text)

        async def _generate_and_play():
            communicate = edge_tts.Communicate(speak_text, voice, rate="+25%")
            audio_data = b""
            async for chunk in communicate.stream():
                if chunk["type"] == "audio":
                    audio_data += chunk["data"]
            
            fp = io.BytesIO(audio_data)
            pygame.mixer.music.load(fp, 'mp3')
            pygame.mixer.music.play()
            
            while pygame.mixer.music.get_busy():
                await asyncio.sleep(0.02)
                
        asyncio.run(_generate_and_play())
    except Exception as e:
        print(f"Speech error: {e}")
    finally:
        last_interaction_time = time.time()  
        watcher_pause_event.clear()  

# 🚀 NEW: AI BUSINESS AUTOMATION (EXCEL TO WEB DATA ENTRY) 🚀
def auto_data_entry_from_excel():
    try:
        watcher_pause_event.set()
        smart_speak("Please select the Excel or CSV file for data entry.", "बॉस, डेटा एंट्रीसाठी एक्सेल किंवा CSV फाईल निवडा.")
        
        root = tk.Tk()
        root.attributes('-topmost', True)
        root.withdraw()
        file_path = filedialog.askopenfilename(title="Select Excel File for Auto Entry", filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv")])
        
        if not file_path:
            smart_speak("File selection cancelled.", "फाईल निवड रद्द केली.")
            watcher_pause_event.clear()
            return False
            
        smart_speak("File loaded. Reading the data...", "फाईल लोड झाली. डेटा वाचत आहे...")
        
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
            
        total_rows = len(df)
        smart_speak(f"Found {total_rows} entries. Please open your form or software, click on the first input box, and DO NOT touch the mouse or keyboard. Starting in 10 seconds...", 
                    f"मला {total_rows} लोकांचा डेटा सापडला आहे. कृपया तुमचे फॉर्म किंवा सॉफ्टवेअर उघडा आणि पहिल्या बॉक्समध्ये क्लिक करून ठेवा. मी १० सेकंदात टायपिंग चालू करेन.")
        
        for i in range(10, 0, -1):
            print(f"⏳ [Automation starting in {i} seconds...]")
            time.sleep(1)
            
        print("🚀 [GHOST TYPING AUTOMATION STARTED]")
        
        for index, row in df.iterrows():
            print(f"➡️ Entering Data for Row {index + 1}...")
            for col_name, item in row.items():
                data_str = str(item)
                if data_str.lower() != 'nan' and data_str.lower() != 'nat' and data_str.strip() != '':
                    pyautogui.write(data_str, interval=0.03)
                
                pyautogui.press('tab')
                time.sleep(0.2)
            
            pyautogui.press('enter')
            time.sleep(2) 
            
        smart_speak("Business automation complete! All data has been successfully entered.", "डेटा एंट्री पूर्ण झाली बॉस! सगळा डेटा ऑटोमॅटिकली भरला आहे.")
        watcher_pause_event.clear()
        return True
    except Exception as e:
        print(f"Data Entry Error: {e}")
        smart_speak("An error occurred during data entry.", "डेटा भरताना काहीतरी अडचण आली.")
        watcher_pause_event.clear()
        return False

# 🚀 ADVANCED AI CODING & BUG FIXING PARTNER 🚀
def debug_and_explain_code():
    global last_ai_response, last_generated_file
    try:
        watcher_pause_event.set()
        smart_speak("Please select the code file you want me to check and fix.", "बॉस, कृपया ती कोड फाईल सिलेक्ट करा ज्यातील चुका मला दुरुस्त करायच्या आहेत.")
        
        root = tk.Tk()
        root.attributes('-topmost', True)
        root.withdraw() 
        file_path = filedialog.askopenfilename(title="Select Code File for Niva to Debug")
        
        if file_path:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                code_content = f.read()
                
            smart_speak("Code loaded! Analyzing for bugs and generating a simple explanation. Please wait...", "कोड वाचला बॉस! मी यातल्या चुका शोधत आहे आणि सोप्या भाषेत समजावून सांगण्याची तयारी करत आहे...")
            
            debug_prompt = f"""
            You are an Expert Senior Software Engineer and AI Coding Partner.
            Here is a code file provided by the user:
            ```
            {code_content[:500000]}
            ```
            YOUR EXCLUSIVE TASKS:
            1. Analyze the code completely. Find any bugs, errors, logic flaws, or inefficiencies.
            2. Fix ALL the bugs and rewrite the FULL CORRECTED CODE. You MUST put the new code inside a single markdown block.
            3. Explain what the original code was supposed to do, what exact bugs you found, and how you fixed them in SIMPLE language.
            4. Reply strictly based on the user's current interaction language constraint.
            """
            
            response_text = ask_ai_safe(debug_prompt)
            last_ai_response = response_text  
            
            smart_speak("Analysis complete. Check the screen for the explanation.", "बॉस, कोड फिक्स झाला आहे! मी काय दुरुस्त केलंय ते ऐका.")
            
            file_status = check_and_open_code_in_notepad(response_text)
            if not isinstance(file_status, str): 
                speak(response_text)
            else:
                last_generated_file = file_status
                speak(response_text)
                speak("The completely fixed code is ready and open. Should I run it?", "नवीन फिक्स केलेला कोड तयार आहे. तो रन करू का?")
        else:
            smart_speak("File selection cancelled.", "काम रद्द केले.")
            
        watcher_pause_event.clear()
        return True
    except Exception as e:
        print(f"Code Analysis Error: {e}")
        smart_speak("Sorry, I could not read this code file.", "माफ करा बॉस, ही फाईल वाचण्यात काहीतरी अडचण आली.")
        watcher_pause_event.clear()
        return False

# 🚀 AUTO PROFESSIONAL PDF INVOICE GENERATOR 🚀
def create_professional_invoice(invoice_data):
    try:
        from fpdf import FPDF
        
        pdf = FPDF()
        pdf.add_page()
        
        biller_name = str(invoice_data.get("biller_name", "MICROSOFT BILL CORPORATION")).upper()
        client_name = invoice_data.get("client_name", "Zayed Supermarkets, Sharjah")
        
        pdf.set_font("Arial", 'B', 18)
        pdf.set_text_color(0, 51, 102) 
        pdf.cell(0, 10, biller_name, ln=True, align='C')
        
        pdf.set_font("Arial", '', 10)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 6, "Authorized Premium Wholesale Supplier & Exporter", ln=True, align='C')
        pdf.set_font("Arial", 'I', 9)
        pdf.cell(0, 6, "Automated Invoice Generated via Niva Systems | Reg No: REG-59", ln=True, align='C')
        pdf.ln(5)
        
        pdf.set_draw_color(0, 51, 102)
        pdf.set_line_width(0.5)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(5)
        
        pdf.set_font("Arial", 'B', 11)
        pdf.set_text_color(0, 0, 0)
        inv_number = f"INV-{datetime.datetime.now().strftime('%Y')}-{int(time.time()) % 10000}"
        
        pdf.cell(0, 6, f"Invoice Number: {inv_number}", ln=True)
        pdf.cell(0, 6, f"Date: {datetime.datetime.now().strftime('%d %B %Y')}", ln=True)
        pdf.cell(0, 6, f"Billed To: {client_name}", ln=True)
        pdf.ln(8)
        
        pdf.set_font("Arial", 'B', 11)
        pdf.set_fill_color(0, 51, 102)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(90, 10, "Description of Goods", border=1, fill=True)
        pdf.cell(30, 10, "Quantity", border=1, align='C', fill=True)
        pdf.cell(35, 10, "Rate/Unit", border=1, align='C', fill=True)
        pdf.cell(35, 10, "Total (USD)", border=1, align='C', fill=True)
        pdf.ln(10)
        
        pdf.set_font("Arial", '', 11)
        pdf.set_text_color(0, 0, 0)
        
        grand_total = 0
        for item in invoice_data.get("items", []):
            desc = str(item.get("description", "Premium Commodities & Logistics Supply"))[:40] 
            qty = float(item.get("qty", 1))
            price = float(item.get("price", 0))
            total = qty * price
            grand_total += total
            
            pdf.cell(90, 10, desc, border=1)
            pdf.cell(30, 10, f"{qty} Units", border=1, align='C')
            pdf.cell(35, 10, f"${price:,.2f}", border=1, align='C')
            pdf.cell(35, 10, f"${total:,.2f}", border=1, align='C')
            pdf.ln(10)
            
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(155, 10, "Total Amount:", border=1, align='R')
        pdf.set_text_color(0, 102, 0)
        pdf.cell(35, 10, f"${grand_total:,.2f}", border=1, align='C')
        
        user_profile = os.environ['USERPROFILE']
        safe_client = re.sub(r'\W+', '_', client_name)
        filename = f"Invoice_{safe_client}_{int(time.time())}.pdf"
        
        desktop_path = os.path.join(user_profile, 'Desktop')
        if not os.path.exists(desktop_path):
            desktop_path = os.getcwd() 
            
        filepath = os.path.join(desktop_path, filename)
        pdf.output(filepath)
        return filepath
        
    except ImportError:
        print("\n❌ [FPDF Library Missing] Please open terminal and run: pip install fpdf")
        return "MISSING_LIBRARY"
    except Exception as e:
        print(f"PDF Error: {e}")
        return None

def generate_invoice_from_command(command_lower):
    watcher_pause_event.set()
    smart_speak("Analyzing billing details. Generating a professional PDF invoice...", "बिलाची माहिती तपासून प्रोफेशनल पीडीएफ (PDF) इनव्हॉइस बनवत आहे. कृपया थांबा...")
    
    extraction_prompt = f"""
    Extract invoice/billing details accurately from this user command: '{command_lower}'.
    Rules:
    1. "biller_name": The company generating the bill. If not explicitly mentioned, default to "Microsoft Bill Corporation".
    2. "client_name": The person or company receiving the bill (e.g., "to Zayed Supermarkets").
    3. "items": A list of dictionaries representing the products/services. Each dictionary must have "description" (string), "qty" (number, default 1), and "price" (number). 
    Return ONLY a raw JSON dictionary without markdown.
    """
    
    response_text = ask_ai_safe(extraction_prompt).replace('```json', '').replace('```', '').strip()
    
    try:
        match = re.search(r'\{.*\}', response_text, re.DOTALL)
        if match:
            invoice_data = json.loads(match.group(0))
        else:
            invoice_data = {"biller_name": "Microsoft Bill Corporation", "client_name": "Valued Customer", "items": [{"description": "Premium Commodities", "qty": 1, "price": 0}]}
            
        result_path = create_professional_invoice(invoice_data)
        
        if result_path == "MISSING_LIBRARY":
            smart_speak("I need the FPDF library to generate PDFs. Please install it.", "बॉस, पीडीएफ बनवण्यासाठी 'fpdf' लायब्ररी इन्स्टॉल करावी लागेल. टर्मिनलमध्ये 'pip install fpdf' टाका.")
        elif result_path:
            smart_speak("Invoice generated successfully! Opening the PDF now.", "इनव्हॉइस तयार झाले बॉस! पीडीएफ फाईल उघडत आहे.")
            os.startfile(result_path)
        else:
            smart_speak("Failed to generate invoice.", "इनव्हॉइस बनवताना काहीतरी एरर आला.")
            
    except Exception as e:
        print(f"Invoice Generation Error: {e}")
        smart_speak("Error analyzing billing data.", "माहिती वाचताना अडचण आली.")
        
    watcher_pause_event.clear()
    return True

# 🚀 NIVA COMPUTER VISION ENGINE (LIVE CAMERA ANALYSIS WITH REAL DATA) 🚀
def analyze_with_camera(user_query):
    try:
        smart_speak("Opening camera. Please hold the object or document in front of the camera for 3 seconds.", "कॅमेरा चालू करत आहे. कृपया वस्तू किंवा फळ कॅमेऱ्यासमोर ३ सेकंद धरून ठेवा.")
        
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            smart_speak("Sorry Boss, I couldn't access the camera.", "माफ करा बॉस, कॅमेरा चालू होऊ शकला नाही. कॅमेरा कनेक्टेड आहे का तपासा.")
            return False
            
        start_time = time.time()
        while time.time() - start_time < 3:
            ret, frame = cap.read()
            if ret:
                cv2.imshow('Niva Vision (Capturing...)', frame)
                cv2.waitKey(1)
                
        ret, frame = cap.read()
        cv2.destroyAllWindows()
        cap.release()
        
        if not ret:
            smart_speak("Failed to capture image.", "इमेज कॅप्चर करता आली नाही.")
            return False
            
        file_name = "niva_vision_temp.jpg"
        cv2.imwrite(file_name, frame)
        smart_speak("Image captured! Let me analyze it carefully...", "फोटो काढला! मी सविस्तर आणि अचूक माहिती तपासत आहे...")
        
        img = PIL.Image.open(file_name)
        vision_prompt = f"""
        Analyze this live image carefully. The user asked: '{user_query}'.
        1. Identify the object PERFECTLY.
        2. Use your Google Search capability to fetch REAL, LIVE, and FACTUAL data about the identified object. DO NOT output fake data.
        3. If it is a fruit like Mosambi, evaluate its visual quality.
        4. Explain it clearly and comprehensively like an expert.
        """
        
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=[img, vision_prompt],
            config=config
        )
        
        answer = response.text
        global last_ai_response
        last_ai_response = answer
        
        print(f"\n👁️ [Vision Analysis Result]:\n{answer}\n")
        speak(answer)
        
        if os.path.exists(file_name):
            os.remove(file_name)
            
        return True
    except Exception as e:
        print(f"Vision Error: {e}")
        smart_speak("There was an error analyzing the image.", "इमेज चेक करताना काहीतरी तांत्रिक अडचण आली.")
        return False

def check_and_open_code_in_notepad(text_content):
    try:
        pattern = r'```([^\n]*)\n([\s\S]*?)\n```'
        matches = re.findall(pattern, text_content)
        saved_file = None
        if matches:
            lang_ext_map = {
                'python': '.py', 'py': '.py', 'html': '.html', 'htm': '.html',
                'css': '.css', 'javascript': '.js', 'js': '.js', 'jsx': '.jsx',
                'json': '.json', 'bash': '.sh', 'sh': '.sh', 'bat': '.bat',
                'cmd': '.bat', 'java': '.java', 'cpp': '.cpp', 'c': '.c'
            }
            
            for first_line, code in matches:
                code_to_write = code.strip()
                if not code_to_write: continue
                lang_line = first_line.strip()
                lang = '' 
                if lang_line:
                    parts = re.split(r'[:\s]', lang_line)
                    if parts: lang = parts[0].lower().strip()
                
                if lang in ['text', 'txt', 'prompt', 'md', '']:
                    continue
                    
                ext = lang_ext_map.get(lang, '.txt')
                prefix = "GFC_Generated_Code" if ext in ['.py', '.html', '.js', '.java', '.cpp'] else "GFC_Generated_Document"
                temp_filename = f"{prefix}_{int(time.time())}{ext}"
                
                with open(temp_filename, "w", encoding="utf-8") as f:
                    f.write(code_to_write)
                    
                subprocess.Popen(['notepad.exe', temp_filename])
                if not saved_file: saved_file = temp_filename
            return saved_file if saved_file else True
    except Exception as e:
        print(f"Notepad Router Error: {e}")
    return False

def analyze_document_with_ai(file_path, user_question):
    try:
        smart_speak("Checking the file, please hold on.", "फाईल चेक करत आहे, कृपया थांबा.")
        file_ext = file_path.lower().split('.')[-1]
        document_text = ""
        uploaded_file = None

        if file_ext in ['xlsx', 'xls']:
            import pandas as pd
            df = pd.read_excel(file_path)
            document_text = df.to_csv(index=False)
        elif file_ext in ['txt', 'csv', 'py', 'html', 'json', 'js', 'css']:
            with open(file_path, 'r', encoding='utf-8') as f:
                document_text = f.read()
        elif file_ext in ['pdf', 'png', 'jpg', 'jpeg']:
            uploaded_file = client.files.upload(file=file_path)
        else:
            return f"Sorry Boss, .{file_ext} format is not supported yet."

        smart_speak("File data loaded! Processing your request.", "फाईल डेटा लोड झाला! माहिती शोधत आहे...")

        niva_instruction = f"""
        User Request: {user_question}
        STRICT INSTRUCTIONS FOR NIVA:
        1. If the user asks to read EVERYTHING or SHOW ALL, extract the entire data clearly. 
        2. Always answer in pure, professional language strictly based on user's query language.
        """

        response_text = ask_ai_safe(f"Document Content:\n{document_text[:500000]}\n\n{niva_instruction}")
        if not response_text and uploaded_file:
            response_text = client.models.generate_content(model='gemini-2.5-flash', contents=[uploaded_file, niva_instruction]).text
            
        return response_text if response_text else "Sorry Boss, server is too busy to process the file."

    except Exception as e:
        print(f"RAG Error: {e}")
        return "Sorry Boss, error while reading the file."

def force_open_in_chrome(url):
    chrome_path = "C:/Program Files/Google/Chrome/Application/chrome.exe"
    if os.path.exists(chrome_path):
        subprocess.Popen([chrome_path, "--profile-directory=Default", url])
    else:
        webbrowser.open(url)

def auto_find_file_or_folder(name_query):
    if not name_query: return ""
    user_profile = os.environ['USERPROFILE']
    search_paths = [
        os.path.join(user_profile, 'Desktop'), 
        os.path.join(user_profile, 'Documents'), 
        os.path.join(user_profile, 'Downloads')
    ]
    
    query_clean = name_query.lower().strip()
    for root_path in search_paths:
        if not os.path.exists(root_path): continue
        try:
            for root, dirs, files in os.walk(root_path):
                for f in files:
                    if query_clean == f.lower() or query_clean in f.lower().split('.')[0]:
                        return os.path.join(root, f)
                for d in dirs:
                    if query_clean == d.lower() or query_clean in d.lower():
                        return os.path.join(root, d)
        except: continue
    return ""

def decode_field(field):
    if not field: return "Unknown"
    decoded_parts = decode_header(field)
    result = ""
    for text, charset in decoded_parts:
        if isinstance(text, bytes):
            result += text.decode(charset or 'utf-8', errors='replace')
        else:
            result += text
    return result

# 🚀 BEAUTIFUL EXCEL GENERATOR & LEAD MINER 🚀
def style_excel_file(filepath):
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center_alignment
                cell.border = thin_border
            
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try: 
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = min((max_length + 3), 50)
            ws.column_dimensions[column].width = adjusted_width
            
        wb.save(filepath)
    except Exception as e:
        print(f"Excel Styling Error: {e}")

def generate_leads_to_excel(user_command):
    try:
        intent_prompt = f"""
        Extract the precise search entity and the exact number requested from this user command: '{user_command}'.
        If no number is specified, default to 10. If the user asks for 5, set count to 5. If 20, set 20.
        Return ONLY a raw JSON dictionary. Format: {{"topic": "...", "count": 10}}
        """
        intent_response = ask_ai_safe(intent_prompt)
        match = re.search(r'\{.*\}', intent_response, re.DOTALL)
        if match:
            intent_data = json.loads(match.group(0))
            topic = intent_data.get("topic", user_command)
            count = intent_data.get("count", 10)
        else:
            topic = user_command
            count = 10
            
        smart_speak(f"Extracting full real details for {count} {topic}. Creating a premium Excel sheet. Please wait...", 
                    f"बॉस, {count} {topic} ची संपूर्ण खरी माहिती गोळा करून एक भारी डिझाईनची एक्सेल शीट बनवत आहे. कृपया थांबा...")
        
        extraction_prompt = f"""
        Perform a live web search to find EXACTLY {count} REAL and active entities for: '{topic}'.
        Find their 'Name', 'Website URL', 'Email Address', 'Phone Number', and 'Physical Address'.
        CRITICAL RULE: DO NOT INVENT OR FAKE DATA. If you cannot find the real Email or Phone on the internet, output "N/A".
        Return the result strictly as a raw JSON list of dictionaries.
        Format: [{{"Name": "Real Corp", "Website": "https://real.com", "Email": "contact@real.com", "Phone": "+12345", "Address": "Dubai, UAE"}}]
        Do not wrap inside markdown code blocks. No explanations. Return ONLY the raw JSON array.
        """
        response_text = ask_ai_safe(extraction_prompt, use_search=True)
        response_text = response_text.replace('```json', '').replace('```', '')
        
        match = re.search(r'\[.*\]', response_text, re.DOTALL)
        if match:
            json_data = match.group(0)
            leads = json.loads(json_data)
            
            if leads:
                df = pd.DataFrame(leads)
                user_profile = os.environ['USERPROFILE']
                filename = f"GFC_Data_Extract_{int(time.time())}.xlsx"
                
                desktop_path = os.path.join(user_profile, 'Desktop')
                if not os.path.exists(desktop_path):
                    desktop_path = os.getcwd() 
                    
                filepath = os.path.join(desktop_path, filename)
                df.to_excel(filepath, index=False)
                
                style_excel_file(filepath)
                
                smart_speak("Extraction complete! The beautifully formatted Excel file is ready.", 
                            f"माहिती गोळा झाली बॉस! एकदम भारी डिझाईन केलेली एक्सेल फाईल तयार आहे.")
                os.startfile(filepath)
                return True
                
        smart_speak("I couldn't extract complete data. The server might be overloaded.", 
                    "बॉस, सर्व्हरवर लोड असल्यामुळे पूर्ण माहिती काढता आली नाही. कृपया पुन्हा प्रयत्न करा.")
        return False
    except Exception as e:
        print(f"Lead Gen Error: {e}")
        smart_speak("An error occurred during data extraction.", "डेटा काढताना काहीतरी तांत्रिक अडचण आली.")
        return False

# 🚀 MASTER CORPORATE EMAIL READER & REPLY SYSTEM 🚀
def check_and_reply_emails():
    try:
        smart_speak("Checking your inbox for today's new emails, Boss.", "बॉस, आजचे नवीन ईमेल चेक करत आहे.")
        
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(SENDER_EMAIL, APP_PASSWORD)
        mail.select("inbox")
        
        today_date = datetime.datetime.now().strftime("%d-%b-%Y")
        status, messages = mail.search(None, f'(UNSEEN SINCE "{today_date}")')
        email_ids = messages[0].split()
        
        if not email_ids:
            smart_speak("You have no new emails today.", "बॉस, आज तुम्हाला एकही नवीन ईमेल आलेला नाही.")
            return True
            
        smart_speak(f"You have {len(email_ids)} new emails today. Scanning for important ones...", f"बॉस, आज {len(email_ids)} नवीन ईमेल आले आहेत. महत्त्वाचे ईमेल शोधत आहे...")
        
        important_emails_found = 0
        
        for e_id in reversed(email_ids[-5:]): 
            status, msg_data = mail.fetch(e_id, "(RFC822)")
            for response_part in msg_data:
                if isinstance(response_part, tuple):
                    msg = email.message_from_bytes(response_part[1])
                    
                    sender_full = decode_field(msg.get("From"))
                    subject = decode_field(msg.get("Subject"))
                    
                    sender_email_match = re.search(r'<([^>]+)>', sender_full)
                    reply_to_email = sender_email_match.group(1) if sender_email_match else sender_full
                    
                    body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == "text/plain":
                                body = part.get_payload(decode=True).decode(errors='ignore')
                                break
                    else:
                        body = msg.get_payload(decode=True).decode(errors='ignore')
                        
                    clean_sender = sender_full.split('<')[0].replace('"', '').strip() if sender_full else "Unknown"
                    subject = subject if subject else "No Subject"
                    
                    filter_prompt = f"""
                    Analyze this email.
                    Sender: {clean_sender}
                    Subject: {subject}
                    Body: {body[:300]}
                    Task 1: Is this a promotional, marketing, newsletter, system alert, or social media (like Pinterest, Facebook, Instagram) email? Answer strictly 'YES' or 'NO'.
                    Task 2: If 'NO', summarize the email in 1 short sentence.
                    Format: YES/NO | Summary
                    """
                    ai_analysis = ask_ai_safe(filter_prompt).strip()
                    
                    if not ai_analysis:
                        ai_analysis = "NO | Summary unavailable due to network load."

                    if ai_analysis.upper().startswith("YES"):
                        mail.store(e_id, '+FLAGS', '\\Seen')
                        print(f"🚫 [Silently Skipped Promotional Email from: {clean_sender}]")
                        continue
                        
                    important_emails_found += 1
                    summary = ai_analysis.split("|", 1)[-1].strip() if "|" in ai_analysis else ai_analysis
                    
                    announcement = f"Important email from {clean_sender}. The subject is {subject}. Here is the summary: {summary}. Should I reply to this email?"
                    announcement_marathi = f"बॉस, {clean_sender} कडून महत्त्वाचा ईमेल आला आहे. विषय आहे '{subject}'. थोडक्यात माहिती: {summary}. मी याला रिप्लाय देऊ का?"
                    
                    smart_speak(announcement, announcement_marathi)
                    
                    user_reply = get_clean_input().strip().lower()
                    
                    if any(word in user_reply for word in ['yes', 'ho', 'de', 'reply', 'kar', 'sure', 'ok']):
                        smart_speak("What should be the core message of your reply?", "ठीक आहे. रिप्लाय मध्ये थोडक्यात काय सांगायचे आहे?")
                        reply_instruction = get_clean_input().strip()
                        
                        if reply_instruction:
                            draft_prompt = f"""
                            You are Ganraj, Boss of 'Ganraj Fruit Company'. 
                            Write a highly professional, corporate reply to this email:
                            Sender: {clean_sender}
                            Subject: {subject}
                            User's instruction for reply: "{reply_instruction}"
                            Write ONLY the email body. Be polite, business-oriented, and sign off as "Ganraj\nM/S Ganraj Fruit Company".
                            """
                            draft_response = ask_ai_safe(draft_prompt).strip()
                            
                            if draft_response:
                                execute_email_automation(reply_to_email, clean_sender, f"Re: {subject}", draft_response)
                                mail.store(e_id, '+FLAGS', '\\Seen')
                            else:
                                smart_speak("Sorry, I could not draft the reply due to server overload.", "माफ करा बॉस, सर्व्हर लोडमुळे रिप्लाय ड्राफ्ट करता आला नाही.")
                        else:
                            smart_speak("Reply cancelled.", "काहीही न सांगितल्यामुळे रिप्लाय कॅन्सल केला.")
                    else:
                        smart_speak("Okay, I will ignore this email for now.", "ठीक आहे बॉस, मी या ईमेलकडे दुर्लक्ष करत आहे.")
                        mail.store(e_id, '+FLAGS', '\\Seen') 
                        
        if important_emails_found == 0:
            smart_speak("No important business emails found today. All promotional emails were skipped.", "बॉस, आजचे महत्त्वाचे ईमेल्स नाहीत. जे प्रमोशनल ईमेल्स होते, ते मी वाचून सोडून दिले आहेत.")
            
        mail.logout()
        return True
        
    except Exception as e:
        print(f"IMAP Error: {e}")
        smart_speak("Sorry, I encountered an error while reading emails.", "माफ करा बॉस, ईमेल वाचताना काहीतरी तांत्रिक अडचण आली.")
        return False

# 🚀 MASTER CORPORATE EMAIL COPYWRITER AI 🚀
def extract_email_details(command_text):
    try:
        extraction_prompt = f"""
        You are an expert Corporate Email Copywriter for 'Ganraj', the Boss of 'Ganraj Fruit Company' (GFC).
        Extract and generate Email task details from this user command: '{command_text}'
        Rules:
        1. "contact": The name of the person or company to send the email to.
        2. "subject": Generate a highly professional, catchy, and concise subject line for the email.
        3. "body": Generate a COMPLETE, HIGHLY PROFESSIONAL, and CORPORATE-LEVEL email body based on the user's short command. Always sign off the email automatically as "Ganraj\nM/S Ganraj Fruit Company".
        Return ONLY a raw JSON dictionary without markdown.
        Format: {{"contact": "", "subject": "", "body": ""}}
        """
        response_text = ask_ai_safe(extraction_prompt)
        match = re.search(r'\{.*\}', response_text, re.DOTALL)
        if match:
            return json.loads(match.group(0))
        return {"contact": "", "subject": "", "body": ""}
    except Exception as e:
        print(f"AI Email Extraction Error: {e}")
        return {"contact": "", "subject": "", "body": ""}

def execute_email_automation(to_email_address, contact_name, subject, body):
    try:
        smart_speak(f"Drafting and sending email to {contact_name}. Please wait.", f"बॉस, {contact_name} ला व्यावसायिक ईमेल तयार करून पाठवत आहे.")
        
        msg = EmailMessage()
        msg.set_content(body)
        msg['Subject'] = subject
        msg['From'] = SENDER_EMAIL
        msg['To'] = to_email_address

        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(SENDER_EMAIL, APP_PASSWORD)
        server.send_message(msg)
        server.quit()
        
        smart_speak("The email was sent successfully.", "ईमेल यशस्वीरित्या पाठवला बॉस!")
        return True
    except Exception as e:
        print(f"Email Error: {e}")
        smart_speak("Failed to send the email. Please check your credentials or network connection.", "माफ करा बॉस, ईमेल पाठवता आला नाही. पासवर्ड किंवा नेटवर्क तपासा.")
        return False

# 🚀 AI WHATSAPP INTENT EXTRACTOR 🚀
def extract_whatsapp_details(command_text):
    try:
        extraction_prompt = f"""
        Extract WhatsApp task details from this command: '{command_text}'
        Rules:
        1. "contact": The name of the person to send the message/file to.
        2. "message": The exact message text to send. If not mentioned, leave empty.
        3. "file_query": The name of the folder, file, or image to send (e.g., 'GFC folder' -> 'GFC'). Leave empty if not mentioned.
        Return ONLY a raw JSON dictionary without markdown.
        Format: {{"contact": "", "message": "", "file_query": ""}}
        """
        response_text = ask_ai_safe(extraction_prompt)
        match = re.search(r'\{.*\}', response_text, re.DOTALL)
        if match:
            return json.loads(match.group(0))
        return {"contact": "", "message": "", "file_query": ""}
    except Exception as e:
        print(f"AI Extraction Error: {e}")
        return {"contact": "", "message": "", "file_query": ""}

def execute_whatsapp_automation(contact_name, message_text="", file_path=""):
    try:
        task_desc = []
        if message_text: task_desc.append("message")
        if file_path: task_desc.append("file")
        task_str = " and ".join(task_desc)
        
        smart_speak(f"Sending {task_str} to {contact_name}. Please do not touch the PC.", 
                    f"बॉस, {contact_name} ला {task_str} पाठवत आहे. कृपया पीसीला हात लावू नका.")
        
        if file_path:
            print(f"📎 [Copying file to clipboard: {file_path}]")
            subprocess.run(['powershell', '-command', f"Set-Clipboard -Path '{file_path}'"])
        
        force_open_in_chrome("https://web.whatsapp.com/")
        print("\n⏳ [Waiting 12 seconds for WhatsApp Web to load...]")
        time.sleep(12)
        
        screen_width, screen_height = pyautogui.size()
        pyautogui.click(screen_width / 2, screen_height / 2)
        time.sleep(1)

        pyautogui.hotkey('ctrl', 'alt', '/')
        time.sleep(2)
        
        print(f"⌨️ [Typing Contact Name: {contact_name}]")
        pyautogui.write(contact_name, interval=0.1)
        time.sleep(2) 
        pyautogui.press('enter')
        time.sleep(2) 
        
        if message_text:
            print(f"⌨️ [Typing Message: {message_text}]")
            pyautogui.write(message_text, interval=0.05)
            time.sleep(1)
            pyautogui.press('enter')
            time.sleep(1)
            
        if file_path:
            print("📎 [Pasting file in WhatsApp...]")
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(4) 
            pyautogui.press('enter')
            time.sleep(1)
            pyautogui.press('enter') 
            time.sleep(2)
        
        smart_speak("WhatsApp task completed successfully.", "व्हॉट्सॲप टास्क यशस्वीरित्या पूर्ण झाला बॉस!")
        return True
    except Exception as e:
        print(f"WhatsApp Error: {e}")
        smart_speak("Sorry, the WhatsApp task failed.", "माफ करा बॉस, व्हॉट्सॲप टास्क पूर्ण होऊ शकला नाही.")
        return False

def close_requested_app(command_lower):
    if any(word in command_lower for word in ['tab', 'chrome', 'browser', 'youtube', 'वेबसाईट', 'टॅब', 'क्रॉम', 'band', 'close']):
        smart_speak("Closing the current tab.", "सध्याचा टॅब बंद करत आहे.")
        pyautogui.hotkey('ctrl', 'w')
        return True
    
    app_targets = {
        ('notepad', 'नोटपॅड'): "notepad.exe",
        ('cmd', 'terminal', 'कमांड प्रॉम्प्ट'): "cmd.exe",
        ('task manager', 'टास्क मॅनेजर'): "taskmgr.exe"
    }
    
    for keys, process_name in app_targets.items():
        if any(word in command_lower for word in keys):
            smart_speak(f"Closing {process_name.replace('.exe', '')}.", f"बॅकग्राउंडमधून {process_name.replace('.exe', '')} बंद करत आहे.")
            os.system(f"taskkill /f /im {process_name}")
            return True
            
    smart_speak("Closing the current window.", "सध्याची विंडो बंद करत आहे.")
    pyautogui.hotkey('alt', 'f4')
    return True

# 🚀 CONTINUOUS LISTENING SYSTEM 🚀
def get_clean_input():
    global voice_mode_active
    r = sr.Recognizer()
    r.dynamic_energy_threshold = False  
    r.energy_threshold = 150            
    r.pause_threshold = 0.5             
    
    print("\n⌨️ [Type command] | 🟢 [Press 'Alt' to Start Mic] | 🔴 [Press 'Ctrl' to Stop Mic]...", end="", flush=True)
    
    user_typed = ""
    while True:
        if keyboard.is_pressed('alt'):
            if not voice_mode_active:
                voice_mode_active = True
                print("\n🎤 [Mic ON - Continuous Mode (Press 'Ctrl' to Stop)]...")
                time.sleep(0.3)
                
        if keyboard.is_pressed('ctrl'):
            if voice_mode_active:
                voice_mode_active = False
                print("\n🔇 [Mic OFF - Stopped Listening]")
                time.sleep(0.3)
                return "" 

        if voice_mode_active:
            with sr.Microphone() as source:
                try:
                    print("\n👂 Listening...", end="")
                    audio = r.listen(source, timeout=3, phrase_time_limit=10)
                    query = r.recognize_google(audio, language='en-IN', show_all=False) 
                    print(f"\nYou (Spoken): {query}")
                    return query
                except sr.WaitTimeoutError:
                    pass 
                except Exception:
                    pass
        else:
            if msvcrt.kbhit():
                char = msvcrt.getwch()
                if char == '\r' or char == '\n':
                    print()
                    return user_typed
                elif char == '\x08': 
                    user_typed = user_typed[:-1]
                    print(f"\r⌨️ Typing: {user_typed}          \r⌨️ Typing: {user_typed}", end="", flush=True)
                else:
                    user_typed += char
                    print(f"\r⌨️ Typing: {user_typed}          \r⌨️ Typing: {user_typed}", end="", flush=True)
            time.sleep(0.05)

def log_to_excel(note_text):
    try:
        file_name = "gfc_accounts.xlsx"
        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        current_time = datetime.datetime.now().strftime("%I:%M %p")
        new_data = pd.DataFrame([[current_date, current_time, note_text]], columns=["Date", "Time", "Note"])
        if os.path.exists(file_name):
            try:
                df = pd.read_excel(file_name)
                df = pd.concat([df, new_data], ignore_index=True)
            except: df = new_data
        else: df = new_data
        df.to_excel(file_name, index=False)
        return True
    except Exception as e:
        print(f"Excel Error: {e}")
        return False

def handle_maps_and_navigation(command_lower):
    home_location = "Bajar Wahegaon, Maharashtra"
    if any(phrase in command_lower for phrase in ['my live location', 'my location', 'maze location', 'location open kar']):
        if not any(nav_word in command_lower for nav_word in ['start live', 'route', 'रूट']):
            smart_speak("Opening Bajar Wahegaon on Google Maps.", "तुमचे लोकेशन गुगल मॅप्सवर उघडत आहे.")
            force_open_in_chrome(f"https://www.google.com/maps/search/?api=1&query={home_location}")
            return True
        
    if 'start live' in command_lower or 'live route' in command_lower or 'navigation' in command_lower:
        destination = command_lower
        for word in ['start live', 'location', 'my location', 'open kar', 'open', 'route', 'navigation', 'kar']:
            destination = destination.replace(word, "")
        if destination.strip():
            smart_speak(f"Starting navigation to {destination.strip()}.", f"{destination.strip()} चा रूट चालू करत आहे.")
            force_open_in_chrome(f"https://www.google.com/maps/dir/?api=1&origin={home_location}&destination={destination.strip()}&travelmode=driving")
            return True

    if 'location' in command_lower or 'map' in command_lower or 'open kar' in command_lower:
        place = command_lower
        for word in ['open', 'location', 'map', 'open kar', 'kar']:
            place = place.replace(word, "")
        if place.strip():
            smart_speak(f"Opening {place.strip()} on Google Maps.", f"गुगल मॅप्सवर {place.strip()} उघडत आहे.")
            force_open_in_chrome(f"https://www.google.com/maps/search/?api=1&query={place.strip()}")
            return True
    return False

def open_windows_search_app(command_lower):
    stop_words = ['open', 'start', 'chalu kar', 'khol']
    app_name = command_lower
    for word in stop_words: app_name = app_name.replace(word, "")
    app_name = app_name.strip()

    if app_name:
        if any(web_word in app_name for web_word in ['instagram', 'facebook', '.com', 'website', 'folder', 'drive', 'lock', 'shutdown', 'restart', 'youtube', 'play', 'video', 'location', 'map']):
            return False
        if app_name in ['terminal', 'cmd', 'command prompt']: app_name = 'cmd'
        smart_speak(f"Opening {app_name}.", f"{app_name} उघडत आहे.")
        pyautogui.press('win')        
        time.sleep(0.5)                
        pyautogui.write(app_name)      
        time.sleep(0.5)                
        pyautogui.press('enter')       
        return True
    return False

def open_any_website(command_lower):
    stop_words = ['open', 'start', 'chalu kar', 'website']
    clean_site = command_lower
    for word in stop_words: clean_site = clean_site.replace(word, "")
    clean_site = clean_site.strip()

    if clean_site:
        if 'youtube' in clean_site:
            smart_speak("Opening YouTube.", "यूट्यूब उघडत आहे.")
            force_open_in_chrome("https://www.youtube.com")
            return True
        if 'instagram' in clean_site:
            smart_speak("Opening Instagram.", "इन्स्टाग्राम उघडत आहे.")
            force_open_in_chrome("https://www.instagram.com")
            return True
        url_to_open = f"https://www.{clean_site}" if (clean_site.endswith('.com') or clean_site.endswith('.in') or clean_site.endswith('.org')) else f"https://www.{clean_site}.com"
        smart_speak(f"Opening {clean_site} in the browser.", f"ब्राउझरमध्ये {clean_site} उघडत आहे.")
        force_open_in_chrome(url_to_open)
        return True
    return False

def open_requested_folder(command_lower):
    clean_folder = command_lower.replace("open folder", "").replace("folder open", "").replace("open", "").replace("folder", "").strip()
    user_profile = os.environ['USERPROFILE']
    
    folder_map = {
        ('downloads', 'download'): os.path.join(user_profile, 'Downloads'),
        ('documents', 'document'): os.path.join(user_profile, 'Documents'),
        ('desktop'): os.path.join(user_profile, 'Desktop'),
        ('videos', 'video'): os.path.join(user_profile, 'Videos'),
        ('c drive'): 'C:\\',
        ('d drive'): 'D:\\',
        ('e drive'): 'E:\\'
    }
    for keys, path in folder_map.items():
        if any(word in clean_folder for word in keys):
            if os.path.exists(path):
                smart_speak(f"Opening {clean_folder}.", f"{clean_folder} फोल्डर उघडत आहे.")
                os.startfile(path)
                return True
                
    if clean_folder:
        smart_speak(f"Searching for {clean_folder} folder...", f"{clean_folder} फोल्डर पीसीमध्ये शोधत आहे...")
        search_paths = [os.path.join(user_profile, 'Desktop'), os.path.join(user_profile, 'Documents'), user_profile, 'D:\\', 'C:\\']
        for root_path in search_paths:
            if os.path.exists(root_path):
                try:
                    for root, dirs, files in os.walk(root_path):
                        for d in dirs:
                            if d.lower() == clean_folder.lower():
                                target_path = os.path.join(root, d)
                                smart_speak("Found it! Opening the folder.", "सापडले! फोल्डर उघडत आहे.")
                                os.startfile(target_path)
                                return True
                except: continue
    smart_speak("Opening File Explorer.", "फाईल एक्सप्लोरर उघडत आहे.")
    os.startfile("explorer.exe")
    return True

def autonomous_watcher():
    global last_interaction_time, silence_level, voice_mode_active, CURRENT_COMMAND_LANG
    while True:
        time.sleep(2)
        if watcher_pause_event.is_set() or voice_mode_active:
            last_interaction_time = time.time()
            continue
            
        elapsed = time.time() - last_interaction_time
        lang_prompt = "English" if CURRENT_COMMAND_LANG == "english" else "Marathi"
        
        if elapsed > 60 and silence_level == 0:
            silence_level = 1
            try:
                with chat_lock:
                    response = chat.send_message(f"[System Note: The user has been silent for 1 minute. Generate a very short proactive check-in in strictly {lang_prompt} asking if they need help.]")
                    if response.text: speak(response.text.strip())
            except: pass
            
        elif elapsed > 180 and silence_level == 1:
            silence_level = 2
            try:
                with chat_lock:
                    response = chat.send_message(f"[System Note: The user has been silent for 3 minutes. Generate a very short 1-sentence message in strictly {lang_prompt} saying you will stay quiet until called upon.]")
                    if response.text: speak(response.text.strip())
            except: pass

def execute_single_command(cmd):
    global last_interaction_time, last_generated_file, last_ai_response, last_whatsapp_contact
    command_lower = cmd.strip().lower()
    
    # 🚀 Clean the input based on user correction history 🚀
    command_lower = command_lower.replace("hai niva", "").replace("hi niva", "").replace("hey niva", "").strip()
    command_lower = command_lower.replace("dnyaneshwar", "").replace("gfc brothers", "").replace("gfc brother", "").replace("6 april", "").strip()

    # 🚀 SHUTDOWN PC LOGIC (UPDATED FOR YOUR EXAMPLES) 🚀
    if any(word in command_lower for word in ['shutdown', 'shatडाऊन', 'band karo', 'zop aali', 'baher chalalo', 'baher chalo', 'sleep', 'pc बंद कर']):
        watcher_pause_event.set()
        smart_speak("Shutting down the system in 10 seconds. Have a good time, Boss!", "१० सेकंदात पीसी बंद करत आहे. चांगली झोप घ्या बॉस, बाय!")
        time.sleep(2)
        os.system("shutdown /s /t 10")
        sys.exit()

    if command_lower in ['exit', 'quit', 'bye gfc']:
        smart_speak("Goodbye, Boss! Take care.", "गुडबाय बॉस! काळजी घ्या.")
        sys.exit()

    # 🚀 FAVORITES INTEGRATION (UPDATED FOR BORED/CARTOON) 🚀
    if any(w in command_lower for w in ['bore', 'boring', 'कंटाळा', 'bore zalo', 'bor hota hai', 'bor hotay', 'favourite cartoon', 'favorite cartoon', 'doremon', 'doraemon']):
        watcher_pause_event.set()
        smart_speak("Playing your favorite cartoon, Doremon, to lift your mood! Enjoy!", "बॉस, तुमचा मूड फ्रेश करण्यासाठी तुमचा आवडता कार्टून 'डोरेमॉन' लावत आहे! एन्जॉय करा!")
        pywhatkit.playonyt("Doremon Cartoon in Hindi")
        watcher_pause_event.clear()
        return True

    if any(w in command_lower for w in ['bandeya', 'bandeya rey bandeya']):
        watcher_pause_event.set()
        smart_speak("Playing your favorite song, Bandeya Rey Bandeya!", "बॉस, तुमचं आवडतं गाणं 'बंदया रे बंदया' लावत आहे!")
        pywhatkit.playonyt("Bandeya Rey Bandeya full song")
        watcher_pause_event.clear()
        return True

    # 🚀 AI BUSINESS AUTOMATION (DATA ENTRY) 🚀
    if any(w in command_lower for w in ['data entry', 'fill form', 'डेटा भर', 'डेटा एंट्री', 'auto fill', 'form fill']):
        auto_data_entry_from_excel()
        return True

    # 🚀 ADVANCED AI CODING & BUG FIXING PARTNER 🚀
    if any(w in command_lower for w in ['debug', 'fix bug', 'explain code', 'कोड फिक्स', 'कोड समजावून', 'bug fix', 'code check']):
        debug_and_explain_code()
        return True

    # 🚀 COMPUTER VISION ENGINE (LIVE CAMERA) 🚀
    if any(w in command_lower for w in ['look at', 'check this', 'camera', 'डोळे', 'हे बघ', 'कॅमेरा', 'scan this', 'photo', 'फोटो']):
        watcher_pause_event.set()
        analyze_with_camera(command_lower)
        watcher_pause_event.clear()
        return True

    # 🚀 AUTO PDF INVOICE GENERATOR 🚀
    if any(w in command_lower for w in ['invoice', 'generate bill', 'create bill', 'बिल बनव', 'इनव्हॉइस', 'पावती', 'bill', 'receipt']):
        generate_invoice_from_command(command_lower)
        return True

    # 🚀 DYNAMIC WEB SCRAPING & EXCEL CREATION 🚀
    if any(w in command_lower for w in ['scrape', 'find', 'lead', 'importers', 'create excel', 'excel sheet', 'कंपन्या शोध', 'डेटा काढ', 'एक्सेल']):
        watcher_pause_event.set()
        if len(command_lower.split()) > 3:
            generate_leads_to_excel(command_lower)
        else:
            smart_speak("What details should I find and add to the Excel sheet?", "तुम्हाला कशाची एक्सेल शीट बनवायची आहे? (उदा. Top 5 engineering colleges in Dubai)")
            search_topic = get_clean_input().strip()
            if search_topic:
                generate_leads_to_excel(search_topic)
            else:
                smart_speak("Task cancelled.", "काम रद्द केले.")
        watcher_pause_event.clear()
        return True

    # 🚀 READ AND REPLY EMAILS FEATURE 🚀
    is_read_cmd = any(w in command_lower for w in ['read', 'check', 'वाच', 'चेक', 'show', 'dakhav'])
    has_email_word = any(w in command_lower for w in ['email', 'mail', 'ईमेल', 'मेल', 'inbox'])
    is_send_cmd = any(w in command_lower for w in ['send', 'write', 'draft', 'पाठव', 'लिही', 'karo'])

    if has_email_word and is_read_cmd and not is_send_cmd:
        watcher_pause_event.set()
        check_and_reply_emails()
        watcher_pause_event.clear()
        return True

    # 🚀 MASTER EMAIL SMART LOGIC (AI INTEGRATED) 🚀
    if has_email_word:
        watcher_pause_event.set()
        details = extract_email_details(command_lower)
        contact_name = details.get("contact", "").strip()
        subject = details.get("subject", "").strip()
        body = details.get("body", "").strip()
        
        if not contact_name:
            smart_speak("Who should I send this email to?", "हा ईमेल कोणाला पाठवायचा आहे? कृपया नाव किंवा ईमेल आयडी सांगा.")
            contact_name = get_clean_input().strip()
            
        if not body:
            smart_speak(f"What should I write in the email to {contact_name}?", f"ठीक आहे. {contact_name} ला ईमेल मध्ये काय लिहायचे आहे?")
            body = get_clean_input().strip()
            if not subject:
                subject = f"Message from GFC Boss to {contact_name}" 
                
        EMAIL_BOOK = {
            "ganraj": "dnyaneshworkale2@gmail.com",
            "kale": "ddkale74@gmail.com",
            "upwork": "client@upwork.com",
            "dubai": "dubai_trader@example.com"
        }
        
        to_email = ""
        if "@" in contact_name:
            to_email = contact_name.replace(" ", "")
        else:
            for key in EMAIL_BOOK:
                if key in contact_name.lower():
                    to_email = EMAIL_BOOK[key]
                    break
                
        if not to_email:
            smart_speak(f"I don't have the email address for {contact_name}. Please type it in the console.", f"माझ्याकडे {contact_name} चा ईमेल आयडी नाही. कृपया टाईप करा.")
            to_email = get_clean_input().strip()
            to_email = to_email.replace(" ", "").lower() 
            
        if to_email and body:
            execute_email_automation(to_email, contact_name, subject, body)
        else:
            smart_speak("Task cancelled. Email details are missing.", "माहिती अपूर्ण असल्यामुळे ईमेल कॅन्सल केला.")
            
        watcher_pause_event.clear()
        return True

    # 🚀 MASTER WHATSAPP SMART LOGIC 🚀
    if 'whatsapp' in command_lower or ('send' in command_lower and any(w in command_lower for w in ['message', 'msg', 'code', 'file', 'photo', 'folder', 'document', 'image'])):
        watcher_pause_event.set()
        details = extract_whatsapp_details(command_lower)
        contact_name = details.get("contact", "").strip()
        msg_text = details.get("message", "").strip()
        file_query = details.get("file_query", "").strip()
        
        if not contact_name and last_whatsapp_contact:
            contact_name = last_whatsapp_contact
            
        if not contact_name:
            smart_speak("Please tell me the contact name.", "कृपया कोणाला पाठवायचे त्याचे नाव सांगा.")
            contact_name = get_clean_input().strip()
            
        if contact_name:
            last_whatsapp_contact = contact_name
            file_path = ""
            
            if "code" in file_query.lower() or "code" in command_lower:
                if last_generated_file and os.path.exists(last_generated_file):
                    file_path = os.path.abspath(last_generated_file)
            
            if file_query and not file_path:
                smart_speak(f"Searching for '{file_query}'...", f"बॉस, '{file_query}' शोधत आहे...")
                file_path = auto_find_file_or_folder(file_query)
                if not file_path:
                    smart_speak("File not found. Please select it manually.", "फाईल सापडली नाही, कृपया स्वतः निवडा.")
                    root = tk.Tk(); root.attributes('-topmost', True); root.withdraw() 
                    file_path = filedialog.askopenfilename()
            
            if not msg_text and not file_path and not file_query:
                smart_speak(f"What is the message for {contact_name}?", f"ठीक आहे. {contact_name} साठी काय मेसेज आहे?")
                msg_text = get_clean_input().strip()
                
            if msg_text or file_path:
                execute_whatsapp_automation(contact_name, message_text=msg_text, file_path=file_path)
            else:
                smart_speak("Task cancelled as no message or file was provided.", "मेसेज किंवा फाईल नसल्यामुळे काम रद्द केले.")
                
        watcher_pause_event.clear()
        return True

    if any(word in command_lower for word in ['पूर्ण वाच', 'sagal vach', 'read full', 'all read', 'सगळं वाच', 'पूर्ण माहिती दे']):
        if last_ai_response:
            watcher_pause_event.set()
            speak(last_ai_response, force_read=True)
            watcher_pause_event.clear()
        else:
            smart_speak("There is nothing new to read, Boss.", "बॉस, वाचण्यासाठी काहीही नवीन नाही.")
        return True

    if last_generated_file:
        if any(word in command_lower for word in ['yes', 'ho', 'run', 'chalav', 'kar', 'ok', 'yes boss', 'rankar', 'haan']):
            watcher_pause_event.set()
            smart_speak("Executing the file now, Boss.", "ठीक आहे बॉस, फाईल रन करत आहे!")
            if last_generated_file.endswith('.py'):
                subprocess.Popen(['cmd.exe', '/c', 'start', 'cmd.exe', '/k', f'python {last_generated_file}'])
            else:
                os.startfile(last_generated_file)
            last_generated_file = None
            watcher_pause_event.clear()
            return True
        elif any(word in command_lower for word in ['no', 'nako', 'nahi', 'cancel', 'rahu de']):
            watcher_pause_event.set()
            smart_speak("Understood. The file will not be executed.", "ठीक आहे बॉस, फाईल रन करत नाहीये.")
            last_generated_file = None
            watcher_pause_event.clear()
            return True
        else:
            last_generated_file = None 

    if any(word in command_lower for word in ['read file', 'document', 'pdf', 'फाईल वाच', 'डॉक्युमेंट', 'file padho']):
        watcher_pause_event.set()
        smart_speak("Sure, please select the document.", "नक्कीच बॉस! कृपया फाईल निवडा.")
        root = tk.Tk(); root.attributes('-topmost', True); root.withdraw() 
        file_path = filedialog.askopenfilename(title="Select Document for Niva")
        if file_path:
            smart_speak("File loaded. What would you like to know?", "फाईल सिलेक्ट झाली! तुम्हाला यातून काय माहिती हवी आहे?")
            question = get_clean_input() 
            if question and question.strip() != "":
                answer = analyze_document_with_ai(file_path, question)
                last_ai_response = answer  
                smart_speak("Here is the information you requested, Boss.", "बॉस, हे घ्या उत्तर!")
                file_status = check_and_open_code_in_notepad(answer)
                if not isinstance(file_status, str): speak(answer)
        watcher_pause_event.clear()
        return True
        
    if any(word in command_lower for word in ['close', 'close it', 'बंद कर', 'mood nahi', 'mood nahin', 'क्लोज कर', 'band karo']):
        watcher_pause_event.set()
        close_requested_app(command_lower)
        watcher_pause_event.clear()
        return True

    if any(word in command_lower for word in ['task manager', 'टास्क मॅनेजर']):
        watcher_pause_event.set()
        smart_speak("Opening Task Manager.", "टास्क मॅनेजर ओपन करत आहे.")
        pyautogui.hotkey('ctrl', 'shift', 'esc')
        watcher_pause_event.clear()
        return True

    if any(action in command_lower for action in ['open', 'start', 'instagram', 'youtube', 'kholo', 'chalu karo']):
        if any(web_keyword in command_lower for web_keyword in ['instagram', 'facebook', 'google', 'github', 'youtube', '.com', '.in', 'website']):
            watcher_pause_event.set()
            open_any_website(command_lower)
            watcher_pause_event.clear()
            return True

    if any(word in command_lower for word in ['play', 'play song', 'video', 'लाव', 'गाणे', 'lagao', 'gaana']):
        search_query = command_lower
        for word in ['play', 'song', 'video', 'लाव', 'open', 'चालू कर', 'lagao', 'chalu karo']: 
            search_query = search_query.replace(word, "")
        if search_query.strip():
            watcher_pause_event.set()
            smart_speak(f"Playing {search_query.strip()} on YouTube.", f"यूट्यूबवर {search_query.strip()} प्ले करत आहे.")
            pywhatkit.playonyt(search_query.strip())
            watcher_pause_event.clear()
        return True

    if any(word in command_lower for word in ['नोट कर', 'लिहून ठेव', 'लिखो', 'हिशोब', 'note down', 'save this']):
        note = command_lower.replace("नोट कर", "").replace("लिहून ठेव", "").replace("लिखो", "").replace("note down", "").replace("save this", "").strip()
        if note:
            watcher_pause_event.set()
            smart_speak("Saving this note to the Excel sheet.", "एक्सल शीटमध्ये सेव्ह करत आहे.")
            log_to_excel(note)
            watcher_pause_event.clear()
        return True

    if any(word in command_lower for word in ['location', 'map', 'live', 'रूट', 'जालना', 'jalna']):
        if handle_maps_and_navigation(command_lower): return True

    if any(word in command_lower for word in ['lock pc', 'pc lock', 'lock karo']):
        watcher_pause_event.set()
        smart_speak("Locking the computer now.", "पीसी लॉक करत आहे बॉस.")
        ctypes.windll.user32.LockWorkStation()
        watcher_pause_event.clear()
        return True

    if any(word in command_lower for word in ['folder', 'drive', 'chatbot']):
        if open_requested_folder(command_lower): return True

    if any(action in command_lower for action in ['open', 'start', 'chalu kar', 'kholo', 'chalu karo']):
        if open_windows_search_app(command_lower): return True

    if any(word in command_lower for word in ['time', 'vele', 'kiti vajle', 'वेळ', 'samay']):
        current_time_str = datetime.datetime.now().strftime('%I:%M %p')
        smart_speak(f"The current time is {current_time_str}.", f"बॉस, आता वेळ {current_time_str} झाली आहे.")
        return True
        
    elif any(word in command_lower for word in ["screenshot"]):
        watcher_pause_event.set()
        pyautogui.screenshot("screenshot.png")
        smart_speak("Screenshot captured and saved.", "स्क्रीनशॉट सेव्ह केला आहे बॉस.")
        watcher_pause_event.clear()
        return True

    return False 

def main():
    global last_interaction_time, silence_level, last_generated_file, last_ai_response
    print("\n🚀 [NIVA AI - SUPREME INTELLIGENT MASTER CORE 2026] 🚀")
    
    welcome_text = "Ohho Boss! Niva is ready. Press 'Alt' to start mic, 'Ctrl' to stop, and 'Esc' to interrupt."
    speak(welcome_text)
    
    watcher_thread = threading.Thread(target=autonomous_watcher, daemon=True)
    watcher_thread.start()
    
    interrupt_thread = threading.Thread(target=interrupt_listener, daemon=True)
    interrupt_thread.start() 
    
    while True:
        full_command = get_clean_input()
        if not full_command or full_command.strip() == "": continue
        
        update_command_lang(full_command)
        
        last_interaction_time = time.time()
        silence_level = 0
        
        if any(w in full_command.lower() for w in ['whatsapp', 'email', 'mail', 'ईमेल', 'scrape', 'importer', 'कंपन्या', 'create excel', 'excel sheet', 'एक्सेल']) or ('send' in full_command.lower() and ('message' in full_command.lower() or 'folder' in full_command.lower() or 'file' in full_command.lower())):
            sub_commands = [full_command]
        else:
            split_pattern = r'\band\b|\bमग\b|\bत्यानंतर\b|\bआणि\b|\bthen\b|\baani\b|\bmag\b|\bun\b|\baur\b|\bphir\b'
            sub_commands = re.split(split_pattern, full_command, flags=re.IGNORECASE)
        
        for cmd in sub_commands:
            cmd = cmd.strip()
            if not cmd: continue
            
            print(f"\n⚡ Executing Task: [{cmd}]")
            is_local_cmd = execute_single_command(cmd)
            
            if not is_local_cmd:
                try:
                    with chat_lock:
                        response_text = ask_ai_safe(cmd, use_search=True)
                    
                    if "START_IMAGE_GEN:" in response_text:
                        img_prompt = response_text.replace("START_IMAGE_GEN:", "").strip()
                        # generate_ai_image(img_prompt) # Uncomment if image generation is active
                    elif response_text:
                        last_ai_response = response_text 
                        file_status = check_and_open_code_in_notepad(response_text)
                        speak(response_text) 
                        
                        if isinstance(file_status, str):
                            watcher_pause_event.set()
                            if file_status.endswith('.py') or file_status.endswith('.html') or file_status.endswith('.json'):
                                smart_speak("The code is ready. Should I run it?", "बॉस, कोड तयार आहे. रन करू का?")
                                last_generated_file = file_status
                            watcher_pause_event.clear()
                except Exception as e:
                    print(f"\n❌ [GEMINI API ERROR]: {e}")
                    watcher_pause_event.set()
                    smart_speak("Network error. I am still ready for local commands.", "नेटवर्क एरर! पण मी लोकल कमांड्ससाठी तयार आहे.")
                    watcher_pause_event.clear()
            
            time.sleep(1) 

if __name__ == "__main__":
    main()